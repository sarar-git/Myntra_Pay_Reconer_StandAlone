"""
utils.py

Excel formatting helper functions.
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

TOTAL_FILL = PatternFill(
    fill_type="solid",
    start_color="D9EAD3",
    end_color="D9EAD3"
)

TOTAL_FONT = Font(
    bold=True
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def format_payment_register(ws):
    """
    Formats the Payment Register worksheet.
    """

    # -------------------------
    # Header
    # -------------------------

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # -------------------------
    # Body
    # -------------------------

    for row in ws.iter_rows(min_row=2):

        for cell in row:
            cell.border = THIN_BORDER

        # Settlement Date
        row[0].number_format = "dd-mmm-yyyy"

        # Payment Amount
        row[2].number_format = '#,##0.00'

    # -------------------------
    # Grand Total
    # -------------------------

    total_row = ws.max_row + 1

    ws[f"B{total_row}"] = "Grand Total"

    ws[f"C{total_row}"] = f"=SUM(C2:C{total_row-1})"

    ws[f"B{total_row}"].font = TOTAL_FONT
    ws[f"C{total_row}"].font = TOTAL_FONT

    ws[f"B{total_row}"].fill = TOTAL_FILL
    ws[f"C{total_row}"].fill = TOTAL_FILL

    ws[f"B{total_row}"].border = THIN_BORDER
    ws[f"C{total_row}"].border = THIN_BORDER

    ws[f"C{total_row}"].number_format = '#,##0.00'

    # -------------------------
    # Freeze
    # -------------------------

    ws.freeze_panes = "A2"

    # -------------------------
    # Filter
    # -------------------------

    ws.auto_filter.ref = ws.dimensions

    # -------------------------
    # Auto Width
    # -------------------------

    for column_cells in ws.columns:

        max_length = 0

        column = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            try:
                value = str(cell.value)

                if len(value) > max_length:
                    max_length = len(value)

            except Exception:
                pass

        ws.column_dimensions[column].width = max_length + 4
